"""v0.2.4 관리자 리포트 / 엑셀 내보내기 테스트. (읽기·출력 전용)

- 권한: MANAGER/ADMIN 만 접근. STAFF/TEAM_LEADER 차단.
- 다운로드: 4종 엑셀 status 200 + xlsx Content-Type.
- 집계: APPROVED 거래만, 입고/출고/순증감/입고금액 계산, CANCELED/PENDING 제외.
- 필터/페이지네이션 무관 전체 반영.
"""

from decimal import Decimal
from io import BytesIO

from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from core.factories import (
    BaseFixtureTestCase,
    approve_initial_count,
    create_item,
    create_managed_item,
    create_supplier,
)
from inventory.exports import XLSX_CONTENT_TYPE
from inventory.models import ItemCategory, TransactionType
from inventory.report_selectors import get_monthly_summary
from inventory.selectors import get_current_stock
from inventory.services import cancel_transaction, create_stock_in, create_stock_out


REPORT_URLS = [
    "inventory:monthly_report",
    "inventory:monthly_report_export",
    "inventory:stock_export",
    "inventory:transaction_export",
    "inventory:inbound_pending_export",
]


class ReportFixture:
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.sup = create_supplier(name="A업체")
        cls.item = create_item("거즈 5x5", category=ItemCategory.MEDICAL_SUPPLY)
        cls.mi = create_managed_item(
            item=cls.item, department=cls.dept_skin, default_supplier=cls.sup,
            minimum_stock=5,
        )
        approve_initial_count(cls.mi, created_by=cls.manager)


class ReportPermissionTest(ReportFixture, BaseFixtureTestCase):
    def test_staff_and_team_leader_blocked(self):
        for user in (self.staff_skin, self.team_leader_skin):
            self.client.force_login(user)
            for name in REPORT_URLS:
                resp = self.client.get(reverse(name))
                self.assertEqual(resp.status_code, 403, f"{user}/{name}")

    def test_manager_and_admin_allowed(self):
        for user in (self.manager, self.admin):
            self.client.force_login(user)
            for name in REPORT_URLS:
                resp = self.client.get(reverse(name))
                self.assertEqual(resp.status_code, 200, f"{user}/{name}")

    def test_staff_sidebar_has_no_report_menu(self):
        self.client.force_login(self.staff_skin)
        resp = self.client.get(reverse("inventory:dashboard"))
        self.assertNotContains(resp, reverse("inventory:monthly_report"))

    def test_manager_sidebar_has_report_menu(self):
        self.client.force_login(self.manager)
        resp = self.client.get(reverse("inventory:dashboard"))
        self.assertContains(resp, reverse("inventory:monthly_report"))


class ExportContentTypeTest(ReportFixture, BaseFixtureTestCase):
    def test_all_exports_are_xlsx(self):
        self.client.force_login(self.manager)
        for name in ("stock_export", "transaction_export",
                     "inbound_pending_export", "monthly_report_export"):
            resp = self.client.get(reverse(f"inventory:{name}"))
            self.assertEqual(resp.status_code, 200)
            self.assertEqual(resp["Content-Type"], XLSX_CONTENT_TYPE)
            self.assertIn("attachment", resp["Content-Disposition"])
            self.assertIn(".xlsx", resp["Content-Disposition"])
            # 실제로 openpyxl 로 열리는 유효한 워크북
            wb = load_workbook(BytesIO(resp.content))
            self.assertGreaterEqual(len(list(wb.active.iter_rows())), 1)  # 최소 헤더


class MonthlySummaryCalcTest(ReportFixture, BaseFixtureTestCase):
    def _setup_activity(self):
        create_stock_in(user=self.staff_skin, managed_item=self.mi, quantity=10,
                        unit_price=Decimal("1000"))
        create_stock_out(user=self.staff_skin, managed_item=self.mi,
                         transaction_type=TransactionType.OUT_USE, quantity=3)

    def test_summary_calculation(self):
        self._setup_activity()
        today = timezone.localdate()
        rows = get_monthly_summary(self.manager, today.replace(day=1), today)
        row = next(r for r in rows if r["item_name"] == self.item.name)
        self.assertEqual(row["in_qty"], Decimal("10"))
        self.assertEqual(row["out_qty"], Decimal("3"))
        self.assertEqual(row["net"], Decimal("7"))
        self.assertEqual(row["in_amount"], Decimal("10000"))  # 10 * 1000

    def test_canceled_and_pending_excluded(self):
        self._setup_activity()
        # 취소될 입고 (APPROVED → CANCELED)
        tx = create_stock_in(user=self.staff_skin, managed_item=self.mi, quantity=5,
                             unit_price=Decimal("2000"))
        cancel_transaction(user=self.manager, transaction_obj=tx, cancel_reason="오입력")
        today = timezone.localdate()
        rows = get_monthly_summary(self.manager, today.replace(day=1), today)
        row = next(r for r in rows if r["item_name"] == self.item.name)
        # 취소분(5)은 제외 → 여전히 10
        self.assertEqual(row["in_qty"], Decimal("10"))
        # 현재고 계산 원칙도 유지 (10 - 3 = 7)
        self.assertEqual(get_current_stock(self.mi), Decimal("7"))

    def test_missing_unit_price_no_error(self):
        # 단가 없는 입고도 오류 없이 집계 (금액 0 취급)
        create_stock_in(user=self.staff_skin, managed_item=self.mi, quantity=4)
        today = timezone.localdate()
        rows = get_monthly_summary(self.manager, today.replace(day=1), today)
        row = next(r for r in rows if r["item_name"] == self.item.name)
        self.assertEqual(row["in_qty"], Decimal("4"))
        self.assertEqual(row["in_amount"], Decimal("0"))


class TransactionExportDataTest(ReportFixture, BaseFixtureTestCase):
    def _rows(self, resp):
        wb = load_workbook(BytesIO(resp.content))
        return list(wb.active.iter_rows(values_only=True))

    def test_export_reflects_type_filter(self):
        create_stock_in(user=self.staff_skin, managed_item=self.mi, quantity=10,
                        unit_price=Decimal("1000"))
        create_stock_out(user=self.staff_skin, managed_item=self.mi,
                         transaction_type=TransactionType.OUT_USE, quantity=2)
        self.client.force_login(self.manager)
        # IN 만 필터 (오늘 기준)
        resp = self.client.get(
            reverse("inventory:transaction_export"),
            {"transaction_type": TransactionType.IN, "range": "today"},
        )
        rows = self._rows(resp)
        data = rows[1:]  # 헤더 제외
        self.assertEqual(len(data), 1)  # IN 1건만
        self.assertEqual(data[0][1], "입고")  # 거래유형 컬럼

    def test_export_full_result_not_paginated(self):
        # 여러 건 생성 → 전체가 다운로드된다 (화면 페이지네이션과 무관)
        create_stock_in(user=self.staff_skin, managed_item=self.mi, quantity=3,
                        unit_price=Decimal("100"))
        for _ in range(3):
            create_stock_out(user=self.staff_skin, managed_item=self.mi,
                             transaction_type=TransactionType.OUT_USE, quantity=1)
        self.client.force_login(self.manager)
        resp = self.client.get(reverse("inventory:transaction_export"), {"range": "today"})
        data = self._rows(resp)[1:]
        # IN 1 + OUT 3 = 4건 (INITIAL_COUNT 는 오늘 승인이지만 거래일자 오늘 → 포함될 수 있음)
        in_out = [r for r in data if r[1] in ("입고", "사용")]
        self.assertEqual(len(in_out), 4)

    def test_numbers_are_numeric_cells(self):
        create_stock_in(user=self.staff_skin, managed_item=self.mi, quantity=1000,
                        unit_price=Decimal("1500"))
        self.client.force_login(self.manager)
        resp = self.client.get(reverse("inventory:transaction_export"), {"range": "today"})
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        # 수량 컬럼(6번째)에 숫자 셀이 존재
        qty_values = [row[5] for row in ws.iter_rows(min_row=2, values_only=True)]
        self.assertIn(1000, qty_values)  # 문자열 "1,000" 이 아니라 숫자 1000


class StockExportDataTest(ReportFixture, BaseFixtureTestCase):
    def test_stock_export_has_item_row(self):
        self.client.force_login(self.manager)
        resp = self.client.get(reverse("inventory:stock_export"))
        wb = load_workbook(BytesIO(resp.content))
        rows = list(wb.active.iter_rows(values_only=True))
        self.assertEqual(rows[0][0], "부서")  # 헤더
        names = [r[1] for r in rows[1:]]
        self.assertIn(self.item.name, names)
