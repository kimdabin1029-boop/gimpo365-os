"""v0.2.5 운영기록 초기화 / 기준정보 점검 테스트.

- reset_operational_data: dry-run 기본, --yes 시 운영기록만 삭제(기준정보 유지), DEBUG=False 차단.
- master_data_checks: command/웹/엑셀 공통 점검 로직.
- 관리자 화면/엑셀 권한(MANAGER 이상) + xlsx 4시트.
"""

from decimal import Decimal
from io import BytesIO, StringIO

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import override_settings
from django.urls import reverse
from openpyxl import load_workbook

from core.factories import (
    BaseFixtureTestCase,
    approve_initial_count,
    create_item,
    create_managed_item,
    create_supplier,
)
from inventory.exports import XLSX_CONTENT_TYPE
from inventory.master_data_checks import run_master_data_checks
from inventory.models import (
    CartItem,
    Item,
    ItemCategory,
    ManagedItem,
    Order,
    OrderItem,
    StockTransaction,
    Supplier,
)
from inventory.order_services import (
    add_to_cart,
    confirm_order,
    create_stock_in_from_order_item,
)
from inventory.selectors import get_current_stock
from inventory.services import create_stock_in, request_initial_count

User = get_user_model()


def _section(result, key):
    return next(s for s in result["sections"] if s["key"] == key)


class ResetOperationalDataTest(BaseFixtureTestCase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.sup = create_supplier(name="A업체")
        cls.item = create_item("거즈 5x5", category=ItemCategory.MEDICAL_SUPPLY)
        cls.mi = create_managed_item(
            item=cls.item, department=cls.dept_skin, default_supplier=cls.sup,
            minimum_stock=5,
        )
        approve_initial_count(cls.mi, created_by=cls.manager)

    def _make_records(self):
        # 입고 + 주문 + 주문서 기반 입고 + 장바구니
        create_stock_in(user=self.staff_skin, managed_item=self.mi, quantity=10,
                        unit_price=Decimal("1000"))
        add_to_cart(user=self.staff_skin, managed_item=self.mi, supplier=self.sup, quantity=5)
        order = confirm_order(user=self.staff_skin)[0]
        oi = order.items.first()
        create_stock_in_from_order_item(
            user=self.staff_skin, order_item=oi, quantity=2,
            unit_price=Decimal("500"), no_expiration=True,
        )
        add_to_cart(user=self.staff_skin, managed_item=self.mi, supplier=self.sup, quantity=3)

    def _run(self, *args):
        out = StringIO()
        call_command("reset_operational_data", *args, stdout=out)
        return out.getvalue()

    def test_dry_run_deletes_nothing(self):
        self._make_records()
        before = StockTransaction.objects.count()
        out = self._run()  # 기본 dry-run
        self.assertIn("DRY-RUN", out)
        self.assertEqual(StockTransaction.objects.count(), before)
        self.assertTrue(Order.objects.exists())
        self.assertTrue(CartItem.objects.exists())

    @override_settings(DEBUG=True)
    def test_yes_deletes_operational_keeps_master(self):
        self._make_records()
        user_n = User.objects.count()
        dept_n = self.dept_skin.__class__.objects.count()
        sup_n = Supplier.objects.count()
        item_n = Item.objects.count()
        mi_n = ManagedItem.objects.count()

        self._run("--yes")

        # 운영기록 삭제
        self.assertEqual(StockTransaction.objects.count(), 0)
        self.assertEqual(Order.objects.count(), 0)
        self.assertEqual(OrderItem.objects.count(), 0)
        self.assertEqual(CartItem.objects.count(), 0)
        # 기준정보 유지
        self.assertEqual(User.objects.count(), user_n)
        self.assertEqual(self.dept_skin.__class__.objects.count(), dept_n)
        self.assertEqual(Supplier.objects.count(), sup_n)
        self.assertEqual(Item.objects.count(), item_n)
        self.assertEqual(ManagedItem.objects.count(), mi_n)

    @override_settings(DEBUG=True)
    def test_initial_count_flow_after_reset(self):
        self._make_records()
        self._run("--yes")
        # 현재고 0 (거래 삭제됨)
        self.assertEqual(get_current_stock(self.mi), Decimal("0"))
        # 최초재고 재입력 → 입고 정상
        request_initial_count(user=self.manager, managed_item=self.mi, quantity=0)
        tx = create_stock_in(user=self.staff_skin, managed_item=self.mi, quantity=7,
                             unit_price=Decimal("1000"))
        self.assertEqual(tx.quantity_delta, Decimal("7"))
        self.assertEqual(get_current_stock(self.mi), Decimal("7"))

    def test_debug_false_blocked_without_allow_production(self):
        self._make_records()
        with override_settings(DEBUG=False):
            with self.assertRaises(CommandError):
                self._run("--yes")
        # 삭제되지 않음
        self.assertTrue(StockTransaction.objects.exists())


class MasterDataChecksTest(BaseFixtureTestCase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        from inventory.models import ItemCategory
        cls.cat = ItemCategory.MEDICAL_SUPPLY
        cls.sup = create_supplier(name="정상공급사")
        cls.sup_inactive = create_supplier(name="비활성공급사", is_active=False)

    def _mi(self, name, **kw):
        item = create_item(name, category=self.cat, specification=kw.pop("spec", "규격"))
        defaults = dict(department=self.dept_skin, default_supplier=self.sup,
                        minimum_stock=5, storage_location="선반A")
        defaults.update(kw)
        return create_managed_item(item=item, **defaults)

    def test_no_default_supplier_detected(self):
        mi = self._mi("공급사없음", default_supplier=None)
        approve_initial_count(mi, created_by=self.manager)
        r = run_master_data_checks()
        self.assertIn(mi, _section(r, "no_default_supplier")["items"])

    def test_bad_min_stock_detected(self):
        mi = self._mi("최소재고0", minimum_stock=0)
        approve_initial_count(mi, created_by=self.manager)
        r = run_master_data_checks()
        self.assertIn(mi, _section(r, "bad_min_stock")["items"])

    def test_no_storage_location_detected(self):
        mi = self._mi("보관위치없음", storage_location="")
        approve_initial_count(mi, created_by=self.manager)
        r = run_master_data_checks()
        self.assertIn(mi, _section(r, "no_storage_location")["items"])

    def test_inactive_supplier_linked_detected(self):
        mi = self._mi("비활성공급연결", default_supplier=self.sup_inactive)
        approve_initial_count(mi, created_by=self.manager)
        r = run_master_data_checks()
        self.assertIn(mi, _section(r, "inactive_supplier_linked")["items"])

    def test_no_initial_count_detected(self):
        mi = self._mi("최초재고없음")  # 최초재고 미승인
        r = run_master_data_checks()
        self.assertIn(mi, _section(r, "no_initial_count")["items"])

    def test_normal_item_not_flagged_as_warning(self):
        mi = self._mi("정상품목")
        approve_initial_count(mi, created_by=self.manager)
        r = run_master_data_checks()
        for s in r["warning_sections"]:
            self.assertNotIn(mi, s["items"], s["key"])

    def test_command_runs(self):
        self._mi("품목1")
        out = StringIO()
        call_command("check_inventory_master_data", stdout=out)
        self.assertIn("기준정보 점검 결과", out.getvalue())


class MasterDataWebTest(BaseFixtureTestCase):
    def test_permission(self):
        for user, code in (
            (self.staff_skin, 403), (self.team_leader_skin, 403),
            (self.manager, 200), (self.admin, 200),
        ):
            self.client.force_login(user)
            resp = self.client.get(reverse("inventory:master_data_check"))
            self.assertEqual(resp.status_code, code, str(user))

    def test_staff_no_menu(self):
        self.client.force_login(self.staff_skin)
        resp = self.client.get(reverse("inventory:dashboard"))
        self.assertNotContains(resp, reverse("inventory:master_data_check"))

    def test_manager_has_menu(self):
        self.client.force_login(self.manager)
        resp = self.client.get(reverse("inventory:dashboard"))
        self.assertContains(resp, reverse("inventory:master_data_check"))


class MasterDataExportTest(BaseFixtureTestCase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        from inventory.models import ItemCategory
        cls.sup = create_supplier(name="A업체")
        cls.item = create_item("거즈", category=ItemCategory.MEDICAL_SUPPLY)
        cls.mi = create_managed_item(
            item=cls.item, department=cls.dept_skin, default_supplier=cls.sup, minimum_stock=5,
        )

    def test_staff_blocked(self):
        self.client.force_login(self.staff_skin)
        resp = self.client.get(reverse("inventory:master_data_check_export"))
        self.assertEqual(resp.status_code, 403)

    def test_manager_downloads_four_sheets(self):
        self.client.force_login(self.manager)
        resp = self.client.get(reverse("inventory:master_data_check_export"))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], XLSX_CONTENT_TYPE)
        self.assertIn(".xlsx", resp["Content-Disposition"])
        wb = load_workbook(BytesIO(resp.content))
        self.assertEqual(
            wb.sheetnames, ["관리품목", "품목", "공급업체", "점검결과"]
        )
        # 관리품목 시트 헤더 확인
        ws = wb["관리품목"]
        self.assertEqual(ws.cell(row=1, column=1).value, "부서")
