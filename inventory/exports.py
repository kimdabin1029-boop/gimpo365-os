"""엑셀(xlsx) 내보내기 헬퍼. (v0.2.4 · 읽기/출력 전용)

- openpyxl 로 워크북을 만들어 HttpResponse 로 즉시 스트리밍한다(서버에 파일 저장 안 함).
- 숫자(Decimal)는 문자열이 아니라 숫자 셀로 저장해 엑셀에서 합계 계산이 가능하게 한다.
"""

from datetime import date, datetime
from decimal import Decimal

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font

XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def dated_filename(prefix: str) -> str:
    """prefix_YYYYMMDD.xlsx (오늘 날짜)."""
    return f"{prefix}_{timezone.localdate():%Y%m%d}.xlsx"


def _norm(value):
    """엑셀 셀 값 정규화. Decimal→float(숫자 셀), tz-aware datetime→naive."""
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        if timezone.is_aware(value):
            return timezone.localtime(value).replace(tzinfo=None)
        return value
    if isinstance(value, date):
        return value
    return value


def _fill_sheet(ws, headers, rows):
    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([_norm(c) for c in row])


def _download_response(wb, filename):
    response = HttpResponse(content_type=XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def xlsx_response(*, filename, sheet_title, headers, rows):
    """헤더 + 행들로 단일 시트 xlsx 를 만들어 다운로드 응답을 돌려준다.

    rows: 반복 가능한 '행'들. 각 행은 셀 값 리스트. (Decimal 은 숫자로 저장)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or "Sheet1")[:31]
    _fill_sheet(ws, headers, rows)
    return _download_response(wb, filename)


def xlsx_multi_sheet_response(*, filename, sheets):
    """여러 시트를 담은 xlsx 다운로드 응답. (v0.2.5)

    sheets: [(sheet_title, headers, rows), ...]
    """
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거 후 지정 시트만 추가
    for title, headers, rows in sheets:
        ws = wb.create_sheet(title=(title or "Sheet")[:31])
        _fill_sheet(ws, headers, rows)
    if not wb.worksheets:  # 안전장치: 최소 1개 시트
        wb.create_sheet(title="Sheet1")
    return _download_response(wb, filename)
